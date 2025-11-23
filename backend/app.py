from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import subprocess
import time
import psutil
import os
import cv2
import numpy as np
from openpyxl import Workbook, load_workbook
from datetime import datetime
import joblib
import warnings
import google.generativeai as genai
import requests
import json
import re
warnings.filterwarnings('ignore')

app = Flask(__name__, static_folder='../frontend', static_url_path='')
CORS(app)

# Use parent directory for uploads/outputs
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
OUTPUT_FOLDER = os.path.join(BASE_DIR, 'outputs')
RESULTS_FILE = os.path.join(BASE_DIR, 'outputs', 'results.xlsx')  # Save in outputs folder
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Carbon intensity (grams CO2 per kWh) - Dynamic based on user location
# Default: BESCOM Bangalore grid (0.71 tCO2/MWh = 710 g/kWh)
# Will be updated via /api/carbon-intensity endpoint using Gemini AI
DEFAULT_CARBON_INTENSITY = 710  # g CO2/kWh (Karnataka, India)
CARBON_INTENSITY_CACHE = {}  # Session cache: {session_id: {intensity, region, timestamp}}

# Gemini API Configuration
# Set your API key as environment variable: export GEMINI_API_KEY="your-key-here"
GEMINI_API_KEY = os.getenv('GEMINI_API_KEY', '')
if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)
    print("✅ Gemini API configured")
else:
    print("⚠️  GEMINI_API_KEY not set - using default Karnataka carbon intensity")

# EC2 Power Calibration Results (measured on 2025-11-20)
# t2.micro instance-specific power consumption
P_IDLE_W = 0.016  # Watts - idle power consumption
P_MAX_W = 28.000  # Watts - max load power consumption

# Load ML models at startup
try:
    CRF_MODEL = joblib.load(os.path.join(BASE_DIR, 'crf_model.pkl'))
    PRESET_MODEL = joblib.load(os.path.join(BASE_DIR, 'preset_model.pkl'))
    print("✅ ML models loaded successfully!")
    ML_AVAILABLE = True
except Exception as e:
    print(f"⚠️  ML models not found: {e}")
    print("   Falling back to rule-based optimization only")
    CRF_MODEL = None
    PRESET_MODEL = None
    ML_AVAILABLE = False

def predict_optimal_preset(complexity, width, height, fps, size_mb):
    """
    Determine optimal FFmpeg preset based on video complexity
    Uses rule-based approach for reliable, interpretable results
    
    Strategy: Optimized for lifecycle energy (encoding + storage + playback)
             Balances encoding speed with file size for long-term efficiency
    """
    # Lifecycle-optimized preset selection
    if complexity < 3.5:
        # Low complexity (talking heads, static scenes)
        return 'faster'  # Quick encode, good compression
    elif complexity < 7.0:
        # Medium complexity (normal videos)
        return 'fast'  # Balanced speed and file size
    else:
        # High complexity (action, sports)
        return 'medium'  # Preserve quality, minimize file size

def predict_ml_settings(video_path):
    """
    Use ML models to predict optimal CRF and preset
    Returns: (crf, preset) or None if ML not available
    """
    if not ML_AVAILABLE:
        return None
    
    try:
        # Extract features
        features = extract_ml_features(video_path)
        
        # Prepare feature vector in correct order
        feature_vector = [[
            features['edge_density'],
            features['motion_score'],
            features['brightness'],
            features['color_variance'],
            features['resolution'],
            features['fps'],
            features['duration']
        ]]
        
        # Predict CRF and preset
        predicted_crf = int(CRF_MODEL.predict(feature_vector)[0])
        predicted_preset = PRESET_MODEL.predict(feature_vector)[0]
        
        print(f"🤖 ML Prediction: CRF {predicted_crf}, Preset {predicted_preset}")
        
        return predicted_crf, predicted_preset
    
    except Exception as e:
        print(f"ML prediction error: {e}")
        return None

def extract_ml_features(video_path):
    """
    Extract ALL 7 features needed for ML model prediction
    Returns dict with: edge_density, motion_score, brightness, color_variance, resolution, fps, duration
    """
    try:
        cap = cv2.VideoCapture(video_path)
        
        # Get video properties
        fps = cap.get(cv2.CAP_PROP_FPS)
        width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        duration = frame_count / fps if fps > 0 else 0
        
        # Sample 1 frame per second (fast sampling)
        sample_interval = int(fps) if fps > 0 else 1
        
        edge_scores = []
        motion_scores = []
        brightness_scores = []
        color_variances = []
        prev_gray = None
        
        frame_idx = 0
        while cap.isOpened():
            ret, frame = cap.read()
            if not ret:
                break
            
            # Only process sampled frames (1 per second)
            if frame_idx % sample_interval == 0:
                # Convert to grayscale
                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                
                # Feature 1: Edge density
                edges = cv2.Canny(gray, 100, 200)
                edge_scores.append(np.mean(edges) / 255.0)
                
                # Feature 2: Motion (frame difference)
                if prev_gray is not None:
                    diff = cv2.absdiff(gray, prev_gray)
                    motion_scores.append(np.mean(diff) / 255.0)
                prev_gray = gray.copy()
                
                # Feature 3: Brightness
                brightness_scores.append(np.mean(gray) / 255.0)
                
                # Feature 4: Color variance
                color_variances.append(np.std(frame) / 255.0)
            
            frame_idx += 1
        
        cap.release()
        
        # Calculate features
        features = {
            'edge_density': np.mean(edge_scores) if edge_scores else 0,
            'motion_score': np.mean(motion_scores) if motion_scores else 0,
            'brightness': np.mean(brightness_scores) if brightness_scores else 0,
            'color_variance': np.mean(color_variances) if color_variances else 0,
            'resolution': width * height,
            'fps': fps,
            'duration': duration
        }
        
        return features
    
    except Exception as e:
        print(f"Error extracting features: {e}")
        # Return default features
        return {
            'edge_density': 0.05,
            'motion_score': 0.05,
            'brightness': 0.5,
            'color_variance': 0.2,
            'resolution': 1920 * 1080,
            'fps': 30,
            'duration': 30
        }

def analyze_video_complexity(video_path):
    """
    Analyze video complexity using OpenCV
    Returns a complexity score from 0-10 based on edge density and motion
    """
    try:
        features = extract_ml_features(video_path)
        
        # Calculate simple complexity score for display
        edge_normalized = min(10, (features['edge_density'] / 0.15) * 10)
        motion_normalized = min(10, (features['motion_score'] / 0.10) * 10)
        
        complexity = (edge_normalized * 0.6 + motion_normalized * 0.4)
        complexity = min(10, max(0, complexity))
        
        print(f"Debug - Edge: {features['edge_density']:.4f}, Motion: {features['motion_score']:.4f}, Complexity: {complexity:.2f}/10")
        
        return round(complexity, 2)
    
    except Exception as e:
        print(f"Error analyzing video: {e}")
        return 5.0

def calculate_energy(duration, cpu_percent):
    """
    Calculate energy consumption in Joules using calibrated EC2 power model
    Energy (J) = Power (W) × Time (s)
    Power = P_idle + (P_max - P_idle) × (CPU% / 100)
    """
    # Use EC2-calibrated power values (linear interpolation between idle and max)
    power_watts = P_IDLE_W + (P_MAX_W - P_IDLE_W) * (cpu_percent / 100)
    energy_joules = power_watts * duration
    return round(energy_joules, 2)

def calculate_co2(energy_joules, carbon_intensity=None):
    """
    Calculate CO2 emissions from energy consumption
    CO2 (g) = Energy (kWh) × Carbon Intensity (g CO2/kWh)
    """
    # Use provided carbon intensity or default to Karnataka
    if carbon_intensity is None:
        carbon_intensity = DEFAULT_CARBON_INTENSITY
    
    # Convert Joules to kWh (1 kWh = 3,600,000 J)
    energy_kwh = energy_joules / 3600000
    co2_grams = energy_kwh * carbon_intensity
    return round(co2_grams, 2)

def save_to_excel(data):
    """
    Save results to Excel file (append if exists, create if not)
    """
    try:
        print(f"Attempting to save to Excel: {RESULTS_FILE}")
        
        # Check if Excel file exists
        if os.path.exists(RESULTS_FILE):
            print("Excel file exists, loading...")
            wb = load_workbook(RESULTS_FILE)
            ws = wb.active
        else:
            # Create new workbook with headers
            print("Creating new Excel file...")
            wb = Workbook()
            ws = wb.active
            ws.append([
                'Timestamp', 'Filename', 'Complexity', 
                'Normal_Energy_J', 'Rule_Energy_J', 'ML_Energy_J',
                'Rule_Savings_%', 'ML_Savings_%',
                'CO2_Normal_g', 'CO2_Rule_g', 'CO2_ML_g'
            ])
        
        # Append new data
        ws.append([
            data['timestamp'],
            data['filename'],
            data['complexity'],
            data['normal_energy'],
            data['rule_energy'],
            data['ml_energy'],
            data['rule_savings_percent'],
            data['ml_savings_percent'],
            data['co2_normal'],
            data['co2_rule'],
            data['co2_ml']
        ])
        
        # Save workbook
        wb.save(RESULTS_FILE)
        print(f"✅ Results saved successfully to {RESULTS_FILE}")
    
    except Exception as e:
        print(f"❌ Error saving to Excel: {e}")
        import traceback
        traceback.print_exc()

@app.route('/api/carbon-intensity', methods=['POST'])
def get_carbon_intensity():
    """
    Get carbon intensity for a geographic location using Gemini AI
    Request: {"latitude": 12.9716, "longitude": 77.5946}
    Response: {"region": "Karnataka, India", "intensity": 710, "source": "gemini"}
    """
    try:
        data = request.get_json()
        lat = data.get('latitude')
        lon = data.get('longitude')
        
        if not lat or not lon:
            return jsonify({
                'region': 'Karnataka, India (default)',
                'intensity': DEFAULT_CARBON_INTENSITY,
                'source': 'default',
                'error': 'Missing coordinates'
            }), 200
        
        # Check if Gemini API is configured
        if not GEMINI_API_KEY:
            print("⚠️  Gemini API key not set, using default")
            return jsonify({
                'region': 'Karnataka, India (default)',
                'intensity': DEFAULT_CARBON_INTENSITY,
                'source': 'default',
                'error': 'Gemini API not configured'
            }), 200
        
        # Reverse geocode coordinates to get location name using free API
        try:
            geocode_url = f"https://nominatim.openstreetmap.org/reverse?lat={lat}&lon={lon}&format=json"
            headers = {'User-Agent': 'GreenAI-VideoTranscoder/1.0'}
            geo_response = requests.get(geocode_url, headers=headers, timeout=5)
            geo_data = geo_response.json()
            
            # Extract region info
            address = geo_data.get('address', {})
            state = address.get('state', '')
            country = address.get('country', '')
            location_name = f"{state}, {country}" if state and country else f"{country}" if country else "Unknown"
            
            print(f"📍 Geocoded: {lat}, {lon} → {location_name}")
        except Exception as e:
            print(f"Geocoding failed: {e}")
            location_name = f"Location at {lat:.4f}, {lon:.4f}"
        
        # Call Gemini API with strict prompt
        try:
            model = genai.GenerativeModel('gemini-2.0-flash')
            
            prompt = f"""You are a precise data lookup assistant for electricity grid carbon intensity.

Location: {location_name}
Coordinates: {lat}, {lon}

Task: Return the OFFICIAL electricity grid carbon intensity for this region.

Rules:
1. Use the most recent official data (2023-2024) from government/energy authority sources
2. Return ONLY a valid JSON object, no markdown, no explanation
3. Carbon intensity must be in grams CO₂ per kWh (g/kWh)
4. If exact regional data unavailable, use national grid average
5. Value must be between 0 and 1500 (realistic range)

Required JSON format:
{{
  "region": "[State/Province], [Country]",
  "intensity": [number],
  "year": "2023" or "2024",
  "source": "[Official source name]"
}}

Example valid response:
{{
  "region": "Karnataka, India",
  "intensity": 710,
  "year": "2024",
  "source": "BESCOM Emission Factor 2023-24"
}}

Return ONLY the JSON object:"""
            
            response = model.generate_content(prompt)
            response_text = response.text.strip()
            
            print(f"🤖 Gemini raw response: {response_text}")
            
            # Extract JSON from response (remove markdown code blocks if present)
            json_match = re.search(r'\{[^}]+\}', response_text, re.DOTALL)
            if json_match:
                json_str = json_match.group(0)
                gemini_data = json.loads(json_str)
                
                region = gemini_data.get('region', location_name)
                intensity = gemini_data.get('intensity')
                source = gemini_data.get('source', 'Gemini AI')
                year = gemini_data.get('year', '2024')
                
                # Validate intensity is reasonable
                if isinstance(intensity, (int, float)) and 0 <= intensity <= 1500:
                    print(f"✅ Carbon intensity: {region} = {intensity} g/kWh ({year}, {source})")
                    return jsonify({
                        'region': region,
                        'intensity': float(intensity),
                        'source': 'gemini',
                        'year': year,
                        'data_source': source
                    }), 200
                else:
                    print(f"⚠️  Invalid intensity value: {intensity}")
                    raise ValueError("Invalid carbon intensity value")
            else:
                print(f"⚠️  No JSON found in Gemini response")
                raise ValueError("No valid JSON in response")
                
        except Exception as e:
            print(f"❌ Gemini API error: {e}")
            return jsonify({
                'region': f'{location_name} (fallback)',
                'intensity': DEFAULT_CARBON_INTENSITY,
                'source': 'default',
                'error': f'Gemini error: {str(e)}'
            }), 200
            
    except Exception as e:
        print(f"❌ Carbon intensity endpoint error: {e}")
        return jsonify({
            'region': 'Karnataka, India (default)',
            'intensity': DEFAULT_CARBON_INTENSITY,
            'source': 'default',
            'error': str(e)
        }), 200

@app.route('/upload', methods=['POST'])
def upload_video():
    print("=" * 80)
    print("🚀 UPLOAD ROUTE CALLED!")
    print("=" * 80)
    
    if 'video' not in request.files:
        return jsonify({'error': 'No video file'}), 400
    
    file = request.files['video']
    print(f"📹 File received: {file.filename}")
    input_path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(input_path)
    print(f"💾 File saved to: {input_path}")
    
    # Extract video metadata
    try:
        cap = cv2.VideoCapture(input_path)
        width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        fps = int(cap.get(cv2.CAP_PROP_FPS))
        cap.release()
        size_mb = os.path.getsize(input_path) / (1024 * 1024)
        
        video_info = {
            'width': width,
            'height': height,
            'fps': fps,
            'size_mb': size_mb
        }
        print(f"📊 Video: {width}x{height} @ {fps}fps, {size_mb:.2f}MB")
    except Exception as e:
        print(f"⚠️ Could not extract metadata: {e}")
        video_info = None
    
    # Step 1: Get input file size
    input_size_mb = os.path.getsize(input_path) / (1024 * 1024)
    
    # Step 2: Analyze video complexity
    print("Analyzing video complexity...")
    complexity = analyze_video_complexity(input_path)
    print(f"Complexity score: {complexity}/10")
    
    # Step 3: Normal transcoding (standard settings)
    normal_output = os.path.join(OUTPUT_FOLDER, 'normal_' + file.filename)
    normal_energy, normal_time, normal_settings = transcode_video(
        input_path, normal_output, 'normal', complexity, video_info
    )
    
    # Step 4: Rule-based Green AI transcoding
    rule_output = os.path.join(OUTPUT_FOLDER, 'rule_' + file.filename)
    rule_energy, rule_time, rule_settings = transcode_video(
        input_path, rule_output, 'rule', complexity, video_info
    )
    
    # Step 5: ML-based Green AI transcoding (if available)
    if ML_AVAILABLE:
        ml_output = os.path.join(OUTPUT_FOLDER, 'ml_' + file.filename)
        ml_energy, ml_time, ml_settings = transcode_video(
            input_path, ml_output, 'ml', complexity, video_info
        )
    else:
        # Fallback to rule-based if ML not available
        ml_output = rule_output
        ml_energy, ml_time, ml_settings = rule_energy, rule_time, rule_settings.copy()
        ml_settings['mode'] = 'ML (Unavailable - Using Rules)'
    
    # Step 6: Get output file sizes and calculate storage savings
    normal_size_mb = os.path.getsize(normal_output) / (1024 * 1024) if os.path.exists(normal_output) else 0
    rule_size_mb = os.path.getsize(rule_output) / (1024 * 1024) if os.path.exists(rule_output) else 0
    ml_size_mb = os.path.getsize(ml_output) / (1024 * 1024) if os.path.exists(ml_output) else 0
    
    rule_storage_saved_mb = normal_size_mb - rule_size_mb
    rule_storage_saved_percent = (rule_storage_saved_mb / normal_size_mb) * 100 if normal_size_mb > 0 else 0
    
    ml_storage_saved_mb = normal_size_mb - ml_size_mb
    ml_storage_saved_percent = (ml_storage_saved_mb / normal_size_mb) * 100 if normal_size_mb > 0 else 0
    
    # Step 7: Get carbon intensity from request (set by frontend after geolocation)
    carbon_intensity = request.form.get('carbon_intensity')
    if carbon_intensity:
        try:
            carbon_intensity = float(carbon_intensity)
            print(f"📍 Using carbon intensity: {carbon_intensity} g/kWh")
        except:
            carbon_intensity = DEFAULT_CARBON_INTENSITY
            print(f"⚠️  Invalid carbon intensity, using default: {DEFAULT_CARBON_INTENSITY} g/kWh")
    else:
        carbon_intensity = DEFAULT_CARBON_INTENSITY
        print(f"ℹ️  No carbon intensity provided, using default: {DEFAULT_CARBON_INTENSITY} g/kWh")
    
    # Step 8: Calculate energy savings and CO2 with dynamic carbon intensity
    rule_savings = normal_energy - rule_energy
    rule_savings_percent = (rule_savings / normal_energy) * 100 if normal_energy > 0 else 0
    
    ml_savings = normal_energy - ml_energy
    ml_savings_percent = (ml_savings / normal_energy) * 100 if normal_energy > 0 else 0
    
    co2_normal = calculate_co2(normal_energy, carbon_intensity)
    co2_rule = calculate_co2(rule_energy, carbon_intensity)
    co2_ml = calculate_co2(ml_energy, carbon_intensity)
    
    co2_rule_saved = co2_normal - co2_rule
    co2_ml_saved = co2_normal - co2_ml
    
    # Step 9: Save to Excel
    excel_data = {
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'filename': file.filename,
        'complexity': complexity,
        'input_size_mb': round(input_size_mb, 2),
        'normal_size_mb': round(normal_size_mb, 2),
        'rule_size_mb': round(rule_size_mb, 2),
        'ml_size_mb': round(ml_size_mb, 2),
        'rule_storage_saved_mb': round(rule_storage_saved_mb, 2),
        'ml_storage_saved_mb': round(ml_storage_saved_mb, 2),
        'normal_energy': normal_energy,
        'rule_energy': rule_energy,
        'ml_energy': ml_energy,
        'rule_savings_percent': round(rule_savings_percent, 2),
        'ml_savings_percent': round(ml_savings_percent, 2),
        'co2_normal': co2_normal,
        'co2_rule': co2_rule,
        'co2_ml': co2_ml,
        'co2_rule_saved': round(co2_rule_saved, 4),
        'co2_ml_saved': round(co2_ml_saved, 4)
    }
    save_to_excel(excel_data)
    
    # Step 10: Return results with all 3 comparisons
    return jsonify({
        'complexity': complexity,
        'input_size_mb': round(input_size_mb, 2),
        'normal_size_mb': round(normal_size_mb, 2),
        'rule_size_mb': round(rule_size_mb, 2),
        'ml_size_mb': round(ml_size_mb, 2),
        'rule_storage_saved_mb': round(rule_storage_saved_mb, 2),
        'rule_storage_saved_percent': round(rule_storage_saved_percent, 2),
        'ml_storage_saved_mb': round(ml_storage_saved_mb, 2),
        'ml_storage_saved_percent': round(ml_storage_saved_percent, 2),
        'normal_energy': normal_energy,
        'normal_time': normal_time,
        'normal_settings': normal_settings,
        'rule_energy': rule_energy,
        'rule_time': rule_time,
        'rule_settings': rule_settings,
        'ml_energy': ml_energy,
        'ml_time': ml_time,
        'ml_settings': ml_settings,
        'rule_savings': round(rule_savings, 2),
        'rule_savings_percent': round(rule_savings_percent, 2),
        'ml_savings': round(ml_savings, 2),
        'ml_savings_percent': round(ml_savings_percent, 2),
        'co2_normal': co2_normal,
        'co2_rule': co2_rule,
        'co2_ml': co2_ml,
        'co2_rule_saved': round(co2_rule_saved, 4),
        'co2_ml_saved': round(co2_ml_saved, 4),
        'ml_available': ML_AVAILABLE,
        'normal_video_url': f'/outputs/normal_{file.filename}',
        'rule_video_url': f'/outputs/rule_{file.filename}',
        'ml_video_url': f'/outputs/ml_{file.filename}'
    })

def transcode_video(input_path, output_path, mode, complexity, video_info=None):
    """
    Transcode video with mode-specific settings
    Mode 'normal': Uses standard quality settings (baseline)
    Mode 'rule': Uses rule-based adaptive settings
    Mode 'ml': Uses ML-predicted settings
    Returns: (energy, duration, settings_dict)
    """
    cpu_percentages = []
    monitoring = {'active': True}  # Use dict for thread-safe flag
    
    # Get video metadata if provided
    width = video_info.get('width', 1920) if video_info else 1920
    height = video_info.get('height', 1080) if video_info else 1080
    fps = video_info.get('fps', 30) if video_info else 30
    size_mb = video_info.get('size_mb', 50) if video_info else 50
    
    # Measure baseline CPU before starting
    baseline_cpu = psutil.cpu_percent(interval=1.0)

    # Optional FFmpeg warm-up: run a very short FFmpeg invocation to initialize codecs
    # This helps exclude FFmpeg startup/initialization time from the measured duration.
    try:
        warmup_cmd = [
            'ffmpeg', '-hide_banner', '-loglevel', 'error',
            '-ss', '0', '-t', '0.1', '-i', input_path,
            '-f', 'null', '-'  # discard output
        ]
        # Run warm-up (do not fail the whole transcoding if warmup errors)
        subprocess.run(warmup_cmd, capture_output=True, text=True, timeout=10)
        print("FFmpeg warm-up completed")
    except Exception as e:
        print(f"FFmpeg warm-up failed (ignored): {e}")

    # Start CPU monitoring in background thread
    def monitor_cpu():
        while monitoring['active']:
            cpu_percentages.append(psutil.cpu_percent(interval=0.2))

    import threading
    monitor_thread = threading.Thread(target=monitor_cpu, daemon=True)

    # Use a high-resolution timer and start it immediately before the actual transcode
    start_time = time.perf_counter()
    monitor_thread.start()
    
    settings = {}  # Track all encoding settings
    
    if mode == 'ml':
        # ML-based: Use trained Random Forest models to predict optimal settings
        ml_prediction = predict_ml_settings(input_path)
        
        if ml_prediction:
            crf, preset = ml_prediction
            crf = str(crf)
        else:
            # Fallback to rule-based if ML fails
            preset = predict_optimal_preset(complexity, width, height, fps, size_mb)
            if complexity < 3.5:
                crf = '28'
            elif complexity < 7.0:
                crf = '26'
            else:
                crf = '24'
        
        settings = {
            'mode': 'ML-Optimized (Random Forest)',
            'preset': preset,
            'crf': crf,
            'codec': 'H.264 (libx264)',
            'threads': 4,
            'optimization': f'ML-predicted: {preset} preset + CRF {crf} based on 7 features from 192 training videos',
            'strategy': 'Machine Learning model trained on optimal encoding settings for quality-size tradeoff'
        }
        
        print(f"🤖 ML Mode: preset={preset}, CRF={crf}")
        
        cmd = [
            'ffmpeg', '-i', input_path,
            '-c:v', 'libx264',
            '-preset', preset,
            '-crf', crf,
            '-threads', '4',
            '-y', output_path
        ]
    
    elif mode == 'rule':
        # Rule-based: Adaptive preset AND CRF selection based on complexity
        preset = predict_optimal_preset(complexity, width, height, fps, size_mb)
        
        # Adaptive CRF: simpler videos can use higher CRF (more compression, smaller file)
        if complexity < 3.5:
            crf = '28'  # Low complexity: aggressive compression
        elif complexity < 7.0:
            crf = '26'  # Medium complexity: moderate compression
        else:
            crf = '24'  # High complexity: preserve quality
        
        settings = {
            'mode': 'Rule-Based Adaptive',
            'preset': preset,
            'crf': crf,
            'codec': 'H.264 (libx264)',
            'threads': 4,
            'optimization': f'Rule-based: {preset} preset + CRF {crf} for complexity {complexity:.1f}/10',
            'strategy': 'Adaptive encoding based on edge detection + motion analysis'
        }
        
        print(f"📏 Rule-based: complexity={complexity:.1f} → preset={preset}, CRF={crf}")
        
        cmd = [
            'ffmpeg', '-i', input_path,
            '-c:v', 'libx264',
            '-preset', preset,
            '-crf', crf,
            '-threads', '4',
            '-y', output_path
        ]
    
    else:
        # Normal: FFmpeg defaults (industry standard baseline)
        preset = 'default'  # FFmpeg chooses medium automatically
        crf = '23'
        
        settings = {
            'mode': 'Standard (Fixed)',
            'preset': 'default (medium)',
            'crf': crf,
            'codec': 'H.264 (libx264)',
            'threads': 4,
            'optimization': 'None (baseline)',
            'strategy': 'FFmpeg default settings - industry standard'
        }
        
        cmd = [
            'ffmpeg', '-i', input_path,
            '-c:v', 'libx264',
            '-crf', '23',
            '-threads', '4',
            '-y', output_path
        ]
    
    # Run FFmpeg transcoding
    result = subprocess.run(cmd, capture_output=True, text=True)
    
    end_time = time.perf_counter()
    
    # Stop monitoring and wait for thread
    monitoring['active'] = False
    time.sleep(0.3)  # Give thread time to finish
    
    # Calculate metrics
    duration = end_time - start_time
    
    # Subtract baseline CPU to get only transcoding CPU usage
    transcoding_cpu_samples = [max(0, cpu - baseline_cpu) for cpu in cpu_percentages]
    avg_cpu = sum(transcoding_cpu_samples) / len(transcoding_cpu_samples) if transcoding_cpu_samples else 50
    
    energy = calculate_energy(duration, avg_cpu)
    
    # Add runtime metrics to settings
    settings['duration'] = f"{duration:.2f}s"
    settings['avg_cpu'] = f"{avg_cpu:.1f}%"
    settings['energy'] = f"{energy}J"
    
    # Get output file size
    if os.path.exists(output_path):
        output_size_mb = os.path.getsize(output_path) / (1024 * 1024)
        settings['output_size'] = f"{output_size_mb:.2f} MB"
    
    print(f"{mode} mode: duration={duration:.2f}s, baseline_cpu={baseline_cpu:.1f}%, avg_cpu={avg_cpu:.1f}%, energy={energy}J, preset={preset}")
    
    return energy, round(duration, 2), settings

@app.route('/')
def index():
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/outputs/<filename>')
def serve_output(filename):
    """Serve transcoded video files"""
    return send_from_directory(OUTPUT_FOLDER, filename)

if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=False, port=5000)
